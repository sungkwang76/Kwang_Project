#bithumb crawiling 

import time
import re
import requests
import openpyxl
from bs4 import BeautifulSoup

url = 'https://www.bithumb.com/' #빗썸 사이트
html = requests.get(url).text

soup = BeautifulSoup(html, 'html.parser')
coins = soup.select('.coin_list tr')

wb = openpyxl.Workbook() #excel 파일 생성
sheet = wb.active
sheet.title = '비트코인 시세'
i = 0
j = 0
while True :
    try:
        i = 0
        sheet.cell(row = 1, column = 2+j).value = time.strftime('%H:%M:%S',time.localtime()) #1행에 검색 시간 기재 
        for coin in coins :   
            name = coin.select_one('td:nth-of-type(1) p a strong').text.strip() #비트코인 항목 검색
            sheet.cell(row = 2+i, column = 1).value = name #1열에 항목 기재
            price = coin.select_one('td:nth-of-type(2) strong').text #비트코인 시세 검색
            sheet.cell(row = 2+i, column = 2+j).value = price # 2열부터 시세 기재
            i+=1
        j+=1
        time.sleep(300) # 5분마다 시세 검색        
    except KeyboardInterrupt:
        print('\n종료하겠습니다.')
        break
wb.save('/Users/sungkwangkim/Developer/Kwang-project/Python/bitcoin/bithumb_' 
+ time.strftime("%Y%m%d")+ '.xlsx')